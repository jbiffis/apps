#!/usr/bin/env python3
"""
Import 2026 SIM Golf League data from Excel into PostgreSQL.
Reads: /tmp/simleague.xlsx
Connects to: 192.168.86.36:5432/app_simgolf_db
"""

import os
import openpyxl
import sys
import psycopg2
import psycopg2.extras
from datetime import date, timedelta

# ─── DB connection ────────────────────────────────────────────────────────────
conn = psycopg2.connect(
    host=os.environ.get('DB_HOST', '192.168.86.36'),
    port=5432,
    dbname=os.environ.get('DB_NAME', 'app_simgolf_db'),
    user=os.environ.get('DB_USER', 'simgolf_usr'),
    password=os.environ['DB_PASS'],
)
conn.autocommit = True
cur = conn.cursor()

def sql(query, params=None):
    try:
        cur.execute(query, params)
    except Exception as e:
        print(f"SQL ERROR: {e}")
        print(f"Query: {query[:300]}")
        sys.exit(1)

def fetch_one(query, params=None):
    sql(query, params)
    row = cur.fetchone()
    return row[0] if row else None

def fetch_all(query, params=None):
    sql(query, params)
    return cur.fetchall()

wb = openpyxl.load_workbook('/tmp/simleague.xlsx', data_only=True)

print("=" * 60)
print("SIM Golf League 2026 Import")
print("=" * 60)

# ─── 1. Update player names ────────────────────────────────────────────────────
print("\n[1] Updating players...")
# Current DB: Shea(1), Sonat(2), Mike(3), Dave(4), Chris(5), Scott(6), Ryan(7), Pete(8)
# Real players: Shea, Sonat, Michael, Cody, Jeremy, Lance, Tim, Lindsay
player_renames = {
    'Mike':  'Michael',
    'Dave':  'Cody',
    'Chris': 'Jeremy',
    'Scott': 'Lance',
    'Ryan':  'Tim',
    'Pete':  'Lindsay',
}
for old, new in player_renames.items():
    sql("UPDATE players SET name = %s WHERE name = %s", (new, old))
    print(f"  {old} → {new}")

# Build player name → id map
rows = fetch_all("SELECT id, name FROM players ORDER BY id")
player_map = {name: pid for pid, name in rows}
print(f"  Players: {player_map}")

# ─── 2. Create season ──────────────────────────────────────────────────────────
print("\n[2] Creating 2026 season...")
sql("INSERT INTO seasons (year, name) VALUES (2026, '2026 Season')")
season_id = fetch_one("SELECT id FROM seasons WHERE year = 2026")
print(f"  Season id: {season_id}")

# Create 3 tournaments
sql("INSERT INTO tournaments (season_id, number, name) VALUES (%s, 1, 'Tournament 1 — Cabot Cliff')", (season_id,))
sql("INSERT INTO tournaments (season_id, number, name) VALUES (%s, 2, 'Tournament 2 — TPC Scottsdale')", (season_id,))
sql("INSERT INTO tournaments (season_id, number, name) VALUES (%s, 3, 'Tournament 3 — Shadow Creek')", (season_id,))

rows = fetch_all("SELECT id, number FROM tournaments WHERE season_id = %s ORDER BY number", (season_id,))
tournament_map = {num: tid for tid, num in rows}
print(f"  Tournaments: {tournament_map}")

# ─── 3. Create courses ─────────────────────────────────────────────────────────
print("\n[3] Creating courses...")

courses_data = {
    'Paynes Valley': {
        'front': [4,3,4,5,3,4,4,5,4],
        'back':  [3,4,4,5,4,4,3,4,5],
        'front_yds': [0]*9,
        'back_yds':  [0]*9,
    },
    'Cabot Cliff': {
        'front': [5,4,4,3,4,3,5,5,3],
        'back':  [5,4,3,4,3,5,3,4,5],
        'front_yds': [0,0,0,153,0,109,0,0,0],
        'back_yds':  [0,0,200,0,0,0,125,0,0],
    },
    'Whispering Pines': {
        'front': [4,5,3,4,3,4,4,5,4],
        'back':  [3,4,5,4,3,4,5,4,4],
        'front_yds': [0,0,144,0,0,0,0,0,0],
        'back_yds':  [184,0,0,0,176,0,0,0,0],
    },
    'Silvertip': {
        'front': [4,5,3,4,4,3,4,5,4],
        'back':  [5,4,4,4,5,3,4,3,4],
        'front_yds': [0,0,159,0,0,170,0,0,0],
        'back_yds':  [0,0,0,0,0,117,0,0,0],
    },
    'TPC Scottsdale': {
        'front': [4,4,5,3,4,4,3,4,4],
        'back':  [4,4,3,5,4,5,3,4,4],
        'front_yds': [0,0,0,147,0,0,0,0,0],
        'back_yds':  [0,0,0,0,0,0,114,0,0],
    },
    'Bigwin': {
        'front': [4,3,5,3,4,4,5,4,4],
        'back':  [4,4,3,4,4,5,4,3,5],
        'front_yds': [0,0,0,131,0,0,0,0,0],
        'back_yds':  [0,0,0,0,0,0,0,159,0],
    },
    'Stonebridge': {
        'front': [4,3,5,4,5,4,3,4,4],
        'back':  [5,3,4,4,4,3,4,3,5],
        'front_yds': [0,156,0,0,0,0,138,0,0],
        'back_yds':  [0,0,142,0,0,0,144,0,141],
    },
    'Shadow Creek': {
        'front': [4,4,4,5,3,4,5,3,4],
        'back':  [4,4,4,3,4,4,5,3,5],
        'front_yds': [0,0,0,0,140,0,0,160,0],
        'back_yds':  [0,0,0,202,0,0,0,140,0],
    },
    'Greystone': {
        'front': [5,3,4,5,4,3,4,4,4],
        'back':  [5,3,4,4,3,4,4,4,5],
        'front_yds': [0,171,0,0,0,155,0,0,0],
        'back_yds':  [0,0,136,0,140,0,0,0,0],
    },
}

course_map = {}
for name, data in courses_data.items():
    sql("INSERT INTO courses (name) VALUES (%s)", (name,))
    cid = fetch_one("SELECT id FROM courses WHERE name = %s", (name,))
    course_map[name] = cid
    # Front 9 (holes 1-9)
    for i, par in enumerate(data['front']):
        yds = data['front_yds'][i]
        sql("INSERT INTO holes (course_id, hole_number, par, yardage) VALUES (%s, %s, %s, %s)",
            (cid, i+1, par, yds))
    # Back 9 (holes 10-18)
    for i, par in enumerate(data['back']):
        yds = data['back_yds'][i]
        sql("INSERT INTO holes (course_id, hole_number, par, yardage) VALUES (%s, %s, %s, %s)",
            (cid, i+10, par, yds))
    print(f"  {name} (id={cid})")

# ─── 4. Parse scores from Excel ───────────────────────────────────────────────
print("\n[4] Parsing Excel data...")

def parse_wk1and2(wb):
    """Parse practice rounds from Wk1and2 2026 sheet."""
    ws = wb['Wk1and2 2026']
    rows = list(ws.iter_rows(values_only=True))

    players_data = {}
    i = 3  # Start after header rows (row 0=header, 1=par, 2=Player)
    while i < len(rows):
        row = rows[i]
        # Stop at the CTP section
        if row[0] in ('Closest Too',):
            break
        if (row[0] and isinstance(row[0], str) and
            row[0] not in ('Adjusted HC', 'Player') and
            not row[0].startswith('Paynes')):
            name = row[0]
            if name in player_map:
                front = [row[j] for j in range(1, 10)]
                back = [row[j] for j in range(12, 21)]
                players_data[name] = {'front': front, 'back': back}
        i += 1

    return players_data

def parse_tournament_sheet(wb, sheet_name):
    """Parse tournament sheet returning list of (course_name, nine, player_scores) per round."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    rounds = []
    i = 0
    while i < len(rows):
        row = rows[i]
        # Detect course header row: first cell is course name, cols 1-9 are hole numbers 1-9
        if (row[0] and isinstance(row[0], str) and
            not row[0].startswith('Player') and
            not row[0].startswith('Tournament') and
            not row[0].startswith('Grand') and
            row[1] == 1 and row[2] == 2):

            course_name = row[0]

            # Skip Par row and Player header row(s)
            j = i + 2
            while j < len(rows) and (rows[j][0] == 'Player' or rows[j][0] == 'Par'):
                j += 1

            front_scores = {}
            back_scores = {}

            while j < len(rows):
                prow = rows[j]
                # Check for next course header (even if unnamed) BEFORE checking pname
                if prow[1] == 1 and prow[2] == 2:
                    break
                pname = prow[0]
                if pname is None or pname == '':
                    j += 1
                    continue
                if not isinstance(pname, str):
                    break
                if pname in ('Adjusted HC', 'Closest Too'):
                    j += 1
                    continue
                if pname in player_map:
                    front = [prow[k] for k in range(1, 10)]
                    # Back nine starts at col 13 (cols: Total=10, HC=11, Score=12, hole10=13...)
                    back = [prow[k] for k in range(13, 22)]
                    front_scores[pname] = front
                    back_scores[pname] = back
                j += 1

            if front_scores:
                rounds.append((course_name, 'front', front_scores))
            # Only add back nine if it has real scores
            if back_scores and any(
                any(v is not None and v != 0 for v in scores)
                for scores in back_scores.values()
            ):
                rounds.append((course_name, 'back', back_scores))

            i = j
        else:
            i += 1

    return rounds

# Parse practice rounds
practice_scores = parse_wk1and2(wb)
print(f"  Practice players: {list(practice_scores.keys())}")

# Parse all tournament rounds
t1_rounds = parse_tournament_sheet(wb, 'Tournament 1 (2026)')
t2_rounds = parse_tournament_sheet(wb, 'Tournament 2 (2026)')
t3_rounds = parse_tournament_sheet(wb, 'Tournament 3 (2026)')
print(f"  T1 rounds: {[(r[0], r[1]) for r in t1_rounds]}")
print(f"  T2 rounds: {[(r[0], r[1]) for r in t2_rounds]}")
print(f"  T3 rounds: {[(r[0], r[1]) for r in t3_rounds]}")

# ─── 5. Create rounds and insert scores ───────────────────────────────────────
print("\n[5] Creating rounds and inserting scores...")

start_date = date(2026, 1, 3)
week_num = 0

def next_date():
    global week_num
    d = start_date + timedelta(weeks=week_num)
    week_num += 1
    return d.isoformat()

COURSE_NAME_FIX = {
    'TPC Scottdale': 'TPC Scottsdale',
    'TPC Scottdale ': 'TPC Scottsdale',
}

def create_round(season_id, tournament_id, round_number, course_name, nine, played_date, is_practice,
                 ctp_hole=None, ctp_yardage=None):
    course_name = COURSE_NAME_FIX.get(course_name, course_name)
    cid = course_map[course_name]
    sql("""INSERT INTO rounds (season_id, tournament_id, round_number, course_id, nine, played_date,
            is_practice, ctp_hole, ctp_yardage, ctp_prize_amount, chip_in_pot)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 20.00, 0.00)""",
        (season_id, tournament_id, round_number, cid, nine, played_date, is_practice, ctp_hole, ctp_yardage))
    rid = fetch_one(
        "SELECT id FROM rounds WHERE season_id=%s AND round_number=%s AND nine=%s AND course_id=%s",
        (season_id, round_number, nine, cid)
    )
    return rid

def normalize_course(name):
    return COURSE_NAME_FIX.get(name, name)

def insert_scores(round_id, nine, player_scores):
    """Insert per-hole scores. nine='front' → holes 1-9, 'back' → holes 10-18."""
    offset = 0 if nine == 'front' else 9
    for pname, hole_scores in player_scores.items():
        if pname not in player_map:
            continue
        pid = player_map[pname]
        for i, strokes in enumerate(hole_scores):
            if strokes is not None and strokes != 0 and isinstance(strokes, (int, float)):
                hnum = i + 1 + offset
                s = int(strokes)
                sql("""INSERT INTO scores (round_id, player_id, hole_number, strokes)
                        VALUES (%s, %s, %s, %s)
                        ON CONFLICT (round_id, player_id, hole_number) DO UPDATE SET strokes = EXCLUDED.strokes""",
                    (round_id, pid, hnum, s))

round_number = 1

# ── Practice rounds ──
print("  Practice rounds (Paynes Valley)...")
rid = create_round(season_id, None, round_number, 'Paynes Valley', 'front', next_date(), True, ctp_hole=2, ctp_yardage=196)
insert_scores(rid, 'front', {k: v['front'] for k, v in practice_scores.items()})
print(f"    Rd{round_number} Paynes Valley front (id={rid})")
round_number += 1

rid = create_round(season_id, None, round_number, 'Paynes Valley', 'back', next_date(), True, ctp_hole=13, ctp_yardage=139)
insert_scores(rid, 'back', {k: v['back'] for k, v in practice_scores.items()})
print(f"    Rd{round_number} Paynes Valley back (id={rid})")
round_number += 1

# ── T1 rounds ──
t1_id = tournament_map[1]
t1_round_ctps = [
    (6, 153),   # Cabot Cliff front: hole 6, 153yrd
    (16, 200),  # Cabot Cliff back: hole 16, 200yrd
    (3, 144),   # Whispering Pines front: hole 3, 144yrd
    (13, 184),  # Whispering Pines back: hole 13, 184yrd
    (3, 159),   # Silvertip front: hole 3, 159yrd
    (15, 117),  # Silvertip back: hole 15, 117yrd
]

print("  T1 rounds...")
for idx, (course_name, nine, player_scores) in enumerate(t1_rounds):
    ctp_h, ctp_y = t1_round_ctps[idx] if idx < len(t1_round_ctps) else (None, None)
    rid = create_round(season_id, t1_id, round_number, course_name, nine, next_date(), False, ctp_h, ctp_y)
    insert_scores(rid, nine, player_scores)
    print(f"    Rd{round_number} {course_name} {nine} (id={rid})")
    round_number += 1

# ── T2 rounds ──
t2_id = tournament_map[2]
t2_round_ctps = [
    (4, 147),   # TPC Scottsdale front: hole 4, 147yrd
    (17, 114),  # TPC Scottsdale back: hole 17, 114yrd
    (4, 131),   # Bigwin front: hole 4, 131yrd
    (17, 159),  # Bigwin back: hole 17, 159yrd
    (2, 156),   # Stonebridge front: hole 2, 156yrd
    (11, 142),  # Stonebridge back: hole 11, 142yrd
]

print("  T2 rounds...")
for idx, (course_name, nine, player_scores) in enumerate(t2_rounds):
    ctp_h, ctp_y = t2_round_ctps[idx] if idx < len(t2_round_ctps) else (None, None)
    rid = create_round(season_id, t2_id, round_number, course_name, nine, next_date(), False, ctp_h, ctp_y)
    insert_scores(rid, nine, player_scores)
    print(f"    Rd{round_number} {course_name} {nine} (id={rid})")
    round_number += 1

# ── T3 rounds ──
t3_id = tournament_map[3]
t3_round_ctps = [
    (5, 140),   # Shadow Creek front: hole 5, 140yrd
    (14, 202),  # Shadow Creek back: hole 14, 202yrd
    (2, 171),   # Greystone front: hole 2, 171yrd
    (15, 140),  # Greystone back: hole 15, 140yrd
]

print("  T3 rounds (played so far)...")
t3_ctp_idx = 0
for course_name, nine, player_scores in t3_rounds:
    # Skip rounds where all scores are 0 (not yet played)
    all_zero = all(
        all(s is None or s == 0 for s in scores)
        for scores in player_scores.values()
    )
    if all_zero:
        print(f"    Skipping unplayed round: {course_name} {nine}")
        t3_ctp_idx += 1
        continue
    ctp_h, ctp_y = t3_round_ctps[t3_ctp_idx] if t3_ctp_idx < len(t3_round_ctps) else (None, None)
    rid = create_round(season_id, t3_id, round_number, course_name, nine, next_date(), False, ctp_h, ctp_y)
    insert_scores(rid, nine, player_scores)
    print(f"    Rd{round_number} {course_name} {nine} (id={rid})")
    round_number += 1
    t3_ctp_idx += 1

# ─── 6. Insert handicaps ──────────────────────────────────────────────────────
print("\n[6] Inserting handicaps...")

handicaps = {
    1: {'Cody': 19, 'Jeremy': 13.5, 'Lance': 13, 'Shea': 6.5, 'Michael': 6, 'Tim': 5.5, 'Sonat': 5, 'Lindsay': 4},
    2: {'Cody': 22, 'Jeremy': 20, 'Lance': 18, 'Shea': 5.5, 'Michael': 8.5, 'Tim': 7, 'Sonat': 7.5, 'Lindsay': 4},
    3: {'Cody': 21, 'Jeremy': 20, 'Lance': 18, 'Shea': 5, 'Michael': 7, 'Tim': 7, 'Sonat': 7, 'Lindsay': 5},
}

for t_num, hcaps in handicaps.items():
    t_id = tournament_map[t_num]
    for pname, hval in hcaps.items():
        if pname not in player_map:
            continue
        pid = player_map[pname]
        sql("""INSERT INTO handicaps (player_id, season_id, tournament_number, value)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (player_id, season_id, tournament_number) DO UPDATE SET value = EXCLUDED.value""",
            (pid, season_id, t_num, hval))
    print(f"  T{t_num}: {hcaps}")

# ─── 7. Insert prize winnings ─────────────────────────────────────────────────
print("\n[7] Inserting prize winnings...")

tournament_prizes = [
    (1, 'Shea',    't1st', 200.00, 'Tournament 1 1st place'),
    (1, 'Tim',     't2nd',  50.00, 'Tournament 1 2nd place'),
    (2, 'Michael', 't1st', 200.00, 'Tournament 2 1st place'),
    (2, 'Jeremy',  't2nd',  50.00, 'Tournament 2 2nd place'),
]

for t_num, pname, ptype, amount, desc in tournament_prizes:
    if pname not in player_map:
        print(f"  SKIP {pname} not in player_map")
        continue
    pid = player_map[pname]
    t_id = tournament_map[t_num]
    sql("""INSERT INTO prize_winnings (player_id, tournament_id, season_id, type, amount, description)
            VALUES (%s, %s, %s, %s, %s, %s)""",
        (pid, t_id, season_id, ptype, amount, desc))
    print(f"  T{t_num} {ptype}: {pname} ${amount}")

print("\n" + "=" * 60)
print("Import complete!")
print("=" * 60)

# Verification
cur.execute("SELECT COUNT(*) FROM scores")
print(f"\n{cur.fetchone()[0]} scores")
cur.execute("SELECT COUNT(*) FROM rounds")
print(f"{cur.fetchone()[0]} rounds")
cur.execute("SELECT COUNT(*) FROM handicaps")
print(f"{cur.fetchone()[0]} handicaps")
cur.execute("SELECT COUNT(*) FROM courses")
print(f"{cur.fetchone()[0]} courses")
cur.execute("SELECT COUNT(*) FROM players")
print(f"{cur.fetchone()[0]} players")

conn.close()
